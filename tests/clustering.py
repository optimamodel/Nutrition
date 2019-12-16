import xlsxwriter as xw
import sciris as sc
import numpy as np
import sklearn.cluster
import distance
import openpyxl as op

stunt_CE_data = sc.odict()
wast_CE_data = sc.odict()
anaem_CE_data = sc.odict()
mort_CE_data = sc.odict()
deets = op.load_workbook('Testing_Global_cost_effectiveness_parallel_reduced_progs.xlsx')
sheet = deets.get_sheet_by_name('Sheet1')
stunt_words = []
stunt_names = []
wast_words = []
wast_names = []
anaem_words = []
anaem_names = []
mort_words = []
mort_names = []
for i in list(range(129)):
    stunt_sub_words = []
    stunt_names.append(sheet.cell(row=i+2, column=20).value)
    for j in list(range(3)):
        stunt_sub_words.append(sheet.cell(row=i+2, column=j+21).value)  # Replace this line
    stunt_clean_words = [x for x in stunt_sub_words if x != None]
    stunt_words.append(", ".join(stunt_clean_words))
    wast_sub_words = []
    wast_names.append(sheet.cell(row=i + 2, column=28).value)
    for j in list(range(2)):
        wast_sub_words.append(sheet.cell(row=i + 2, column=j + 29).value)  # Replace this line
    wast_clean_words = [x for x in wast_sub_words if x != None]
    wast_words.append(", ".join(wast_clean_words))
    anaem_sub_words = []
    anaem_names.append(sheet.cell(row=i + 2, column=34).value)
    for j in list(range(3)):
        anaem_sub_words.append(sheet.cell(row=i + 2, column=j + 35).value)  # Replace this line
    anaem_clean_words = [x for x in anaem_sub_words if x != None]
    anaem_words.append(", ".join(anaem_clean_words))
    mort_sub_words = []
    mort_names.append(sheet.cell(row=i + 2, column=42).value)
    for j in list(range(6)):
        mort_sub_words.append(sheet.cell(row=i + 2, column=j + 43).value)  # Replace this line
    mort_clean_words = [x for x in mort_sub_words if x != None]
    mort_words.append(", ".join(mort_clean_words))
stunt_words = np.asarray(stunt_words)  # So that indexing with a list will work
stunt_names = np.asarray(stunt_names)
wast_words = np.asarray(wast_words)  # So that indexing with a list will work
wast_names = np.asarray(wast_names)
anaem_words = np.asarray(anaem_words)  # So that indexing with a list will work
anaem_names = np.asarray(anaem_names)
mort_words = np.asarray(mort_words)  # So that indexing with a list will work
mort_names = np.asarray(mort_names)
for i, item in enumerate(stunt_names):
    stunt_CE_data.append(item, stunt_words[i])
for i, item in enumerate(wast_names):
    wast_CE_data.append(item, wast_words[i])
for i, item in enumerate(anaem_names):
    anaem_CE_data.append(item, anaem_words[i])
for i, item in enumerate(mort_names):
    mort_CE_data.append(item, mort_words[i])
stunt_lev_similarity = -1 * np.array([[distance.jaccard(w1, w2) for w1 in stunt_words] for w2 in stunt_words])
wast_lev_similarity = -1 * np.array([[distance.jaccard(w1, w2) for w1 in wast_words] for w2 in wast_words])
anaem_lev_similarity = -1 * np.array([[distance.jaccard(w1, w2) for w1 in anaem_words] for w2 in anaem_words])
mort_lev_similarity = -1 * np.array([[distance.jaccard(w1, w2) for w1 in mort_words] for w2 in mort_words])
stunt_cluster_choices = sc.odict()
wast_cluster_choices = sc.odict()
anaem_cluster_choices = sc.odict()
mort_cluster_choices = sc.odict()
affprop = sklearn.cluster.AffinityPropagation(affinity="precomputed", damping=0.8, max_iter=1000)
affprop.fit(stunt_lev_similarity)
for cluster_id in np.unique(affprop.labels_):
    exemplar = stunt_words[affprop.cluster_centers_indices_[cluster_id]]
    word_cluster = np.unique(stunt_words[np.nonzero(affprop.labels_ == cluster_id)])
    name_cluster = np.unique(stunt_names[np.nonzero(affprop.labels_ == cluster_id)])
    cluster_str = ", ".join(name_cluster)
    stunt_cluster_choices.append(exemplar, cluster_str.split(', '))
print(stunt_cluster_choices)
affprop = sklearn.cluster.AffinityPropagation(affinity="precomputed", damping=0.5, max_iter=1000)
affprop.fit(wast_lev_similarity)
for cluster_id in np.unique(affprop.labels_):
    exemplar = wast_words[affprop.cluster_centers_indices_[cluster_id]]
    word_cluster = np.unique(wast_words[np.nonzero(affprop.labels_ == cluster_id)])
    name_cluster = np.unique(wast_names[np.nonzero(affprop.labels_ == cluster_id)])
    cluster_str = ", ".join(name_cluster)
    wast_cluster_choices.append(exemplar, cluster_str.split(', '))
print(wast_cluster_choices)
affprop = sklearn.cluster.AffinityPropagation(affinity="precomputed", damping=0.5, max_iter=1000)
affprop.fit(anaem_lev_similarity)
for cluster_id in np.unique(affprop.labels_):
    exemplar = anaem_words[affprop.cluster_centers_indices_[cluster_id]]
    word_cluster = np.unique(anaem_words[np.nonzero(affprop.labels_ == cluster_id)])
    name_cluster = np.unique(anaem_names[np.nonzero(affprop.labels_ == cluster_id)])
    cluster_str = ", ".join(name_cluster)
    anaem_cluster_choices.append(exemplar, cluster_str.split(', '))
print(anaem_cluster_choices)
affprop = sklearn.cluster.AffinityPropagation(affinity="precomputed", damping=0.8, max_iter=1000)
affprop.fit(mort_lev_similarity)
for cluster_id in np.unique(affprop.labels_):
    exemplar = mort_words[affprop.cluster_centers_indices_[cluster_id]]
    word_cluster = np.unique(mort_words[np.nonzero(affprop.labels_ == cluster_id)])
    name_cluster = np.unique(mort_names[np.nonzero(affprop.labels_ == cluster_id)])
    cluster_str = ", ".join(name_cluster)
    mort_cluster_choices.append(exemplar, cluster_str.split(', '))
print(mort_cluster_choices)

CE_book = xw.Workbook('Testing_Global_cost_effectiveness_cluster_reduced_progs.xlsx')
CE_sheet = CE_book.add_worksheet()
row = 0
column = 0
CE_sheet.write(row, column, 'stunting')
column += 1
row += 1
for choice in stunt_cluster_choices.keys():
    CE_sheet.write(row, column, choice)
    column += 1
    CE_sheet.write_row(row, column, stunt_cluster_choices[choice])
    column -= 1
    row += 2
column -= 1
CE_sheet.write(row, column, 'wasting')
column += 1
row += 1
for choice in wast_cluster_choices.keys():
    CE_sheet.write(row, column, choice)
    column += 1
    CE_sheet.write_row(row, column, wast_cluster_choices[choice])
    column -= 1
    row += 2
column -= 1
CE_sheet.write(row, column, 'anaemia')
column += 1
row += 1
for choice in anaem_cluster_choices.keys():
    CE_sheet.write(row, column, choice)
    column += 1
    CE_sheet.write_row(row, column, anaem_cluster_choices[choice])
    column -= 1
    row += 2
column -= 1
CE_sheet.write(row, column, 'mortality')
column += 1
row += 1
for choice in mort_cluster_choices.keys():
    CE_sheet.write(row, column, choice)
    column += 1
    CE_sheet.write_row(row, column, mort_cluster_choices[choice])
    column -= 1
    row += 2
CE_book.close()
print('Finished!')







